from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "三国官职数值"

headers = [
    "Id","Name","troopsLimit","cost","level","meritNeeds",
    "commandAdd","strengthAdd","intelligenceAdd","politicsAdd","glamourAdd",
    "addSkills","addFeatures","effect_desc",
    "commandNeed","strengthNeed","intelligenceNeed","politicsNeed","glamourNeed","levelNeed"
]
ws.append(headers)

# 基础数值档位（级别由低到高：14→0）
# level 14=最低, level 0=最高
BASE = {
    # level: (troopsLimit, cost, meritNeeds)
    # troopsLimit 最低5000，最高15000
    # meritNeeds = 晋升到下一级官职所需功绩，最高级(0)为0
    14: (5000,  20,  1000),
    13: (5400,  28,  1500),
    12: (5800,  36,  2200),
    11: (6300,  46,  3300),
    10: (6800,  58,  5000),
    9:  (7400,  72,  7000),
    8:  (8000,  88,  9500),
    7:  (8700,  106, 12500),
    6:  (9500,  126, 16500),
    5:  (10400, 150, 21000),
    4:  (11300, 176, 26500),
    3:  (12200, 205, 33000),
    2:  (13100, 238, 41000),
    1:  (14000, 275, 50000),
    0:  (15000, 320, 0),
}

def attr_add(level, is_military):
    """根据级别和文/武返回属性加成 dict"""
    lv = 14 - level  # 0是最低, 14是最高
    base_val = max(0, min(10, lv // 2 + 1))
    if is_military:
        return {
            "commandAdd": min(10, base_val + (1 if lv >= 8 else 0)),
            "strengthAdd": min(10, base_val + (2 if lv >= 6 else 0)),
            "intelligenceAdd": min(10, base_val - (1 if lv < 8 else 0)),
            "politicsAdd": max(0, base_val - 2),
            "glamourAdd": max(0, base_val - 1),
        }
    else:  # civil
        return {
            "commandAdd": max(0, base_val - 2),
            "strengthAdd": max(0, base_val - 3),
            "intelligenceAdd": min(10, base_val + 2),
            "politicsAdd": min(10, base_val + 3),
            "glamourAdd": min(10, base_val + 1),
        }

def req_attr(level, is_military):
    """根据级别返回需求属性，魅力需求默认均为0（仅特殊官职通过adj设置）"""
    lv = 14 - level  # lv=0(Level14最低) ~ lv=14(Level0最高)
    base = max(0, lv * 5 - 10)
    # levelNeed: 5(Level14最低) ~ 50(Level0最高)
    level_need = max(5, min(50, 5 + lv * 45 // 14))
    if is_military:
        return {
            "commandNeed": min(95, base + 20 + lv * 3),
            "strengthNeed": min(95, base + 25 + lv * 3),
            "intelligenceNeed": max(0, base - 5),
            "politicsNeed": max(0, base - 10),
            "glamourNeed": 0,
            "levelNeed": level_need,
        }
    else:
        return {
            "commandNeed": max(0, base - 5),
            "strengthNeed": max(0, base - 10),
            "intelligenceNeed": min(95, base + 25 + lv * 3),
            "politicsNeed": min(95, base + 30 + lv * 3),
            "glamourNeed": 0,
            "levelNeed": level_need,
        }

def build_attr_desc(a):
    """根据属性加成字典生成描述文本，如 '统率+5，武力+2，智力+8'"""
    parts = []
    attr_names = [
        ("commandAdd", "统率"),
        ("strengthAdd", "武力"),
        ("intelligenceAdd", "智力"),
        ("politicsAdd", "政治"),
        ("glamourAdd", "魅力"),
    ]
    for key, cn_name in attr_names:
        val = a.get(key, 0)
        if val > 0:
            parts.append(f"{cn_name}+{val}")
    if parts:
        return "，".join(parts)
    return ""

def make_office(oid, name, level, is_military, desc, skill="", feature="", adj=None):
    """构造一行官职数据"""
    b = BASE[level]
    a = attr_add(level, is_military)
    r = req_attr(level, is_military)
    if adj:
        a.update(adj.get("add", {}))
        r.update(adj.get("req", {}))
    # 将属性加成自动追加到描述末尾
    attr_str = build_attr_desc(a)
    full_desc = f"{desc}（{attr_str}）" if attr_str else desc
    return (
        oid, name,
        b[0], b[1], level, b[2],
        a["commandAdd"], a["strengthAdd"], a["intelligenceAdd"], a["politicsAdd"], a["glamourAdd"],
        skill, feature, full_desc,
        r["commandNeed"], r["strengthNeed"], r["intelligenceNeed"], r["politicsNeed"], r["glamourNeed"], r["levelNeed"],
    )

# ==================== 官职定义（从低到高：level 14 → 0）====================

oid = 0
rows = []

def add_civil(name, level, desc, **kw):
    global oid
    oid += 1
    rows.append(make_office(oid, name, level, False, desc, **kw))

def add_military(name, level, desc, **kw):
    global oid
    oid += 1
    rows.append(make_office(oid, name, level, True, desc, **kw))

# ---- Level 14 基层吏员 / 基层军士 ----
add_civil("书佐",  14, "抄写文书，县府最低吏员")
add_civil("亭长",  14, "掌一亭治安，缉捕盗贼，汉高祖曾任此职")
add_civil("里魁",  14, "掌一里民事，百户之长")
add_civil("门亭长",14, "守县府门禁，传禀通报")
add_civil("啬夫",  14, "掌一乡赋税诉讼，汉之乡官")
add_military("伍长", 14, "统领五人之长，兵制最基层")
add_military("什长", 14, "统领十人之长，辖二伍")
add_military("斥候", 14, "侦察敌情地形，传递军情")

# ---- Level 13 县吏 / 队屯级军官 ----
add_civil("主记室史",13, "掌县府文书起草记录")
add_civil("令史",    13, "县府文书吏，辅助主记")
add_civil("功曹史",  13, "掌县吏考课与选署")
add_civil("户曹史",  13, "掌一县户籍赋税徭役")
add_civil("贼曹史",  13, "掌一县缉捕盗贼治安")
add_military("队率", 13, "统领一队，约五十人")
add_military("屯长", 13, "统领一屯，约百人")
add_military("游徼", 13, "巡查乡里，缉捕奸盗")

# ---- Level 12 县佐 / 曲级军官 ----
add_civil("县丞",    12, "县令之副贰，佐理一县政务")
add_civil("主簿",    12, "掌一县文书机要，亲近之职")
add_civil("功曹掾",  12, "铨选县吏，考核功过，县之重职")
add_civil("仓曹掾",  12, "掌粮仓出入及仓储管理")
add_civil("金曹掾",  12, "掌钱谷收支市租赋税")
add_military("军候",  12, "统领一曲，约五百人")
add_military("假候",  12, "军候之副贰，佐领一曲")
add_military("军正",  12, "掌军中法令监察")

# ---- Level 11 县令级 / 都尉级 ----
add_civil("县令",    11, "万户以上大县之长，秩千石", adj={"req":{"glamourNeed":18}})
add_civil("县长",    11, "万户以下小县之长，秩四百石", adj={"req":{"glamourNeed":12}})
add_civil("督邮",    11, "郡府派驻监察诸县之官")
add_civil("五官掾",  11, "掌祭祀及诸曹统筹")
add_civil("文学掾",  11, "掌教化学校及儒学")
add_military("都尉", 11, "掌一郡军事，秩比二千石，汉之郡尉")
add_military("别部司马",11, "别领一营之军司马，独立统兵")
add_military("骑司马", 11, "掌骑兵之军司马")

# ---- Level 10 郡佐 / 校尉级 ----
add_civil("郡丞",     10, "郡守副贰，总揽郡府庶务")
add_civil("郡功曹",   10, "主郡吏铨选考核，郡之右职")
add_civil("郡主簿",   10, "掌郡府文书机要及印信")
add_civil("治中从事", 10, "州刺史属官，居中治事")
add_civil("议曹从事", 10, "参议政事，备顾问应对")
add_military("校尉",  10, "掌一营兵马，约千人，秩比二千石")
add_military("军司马",10, "校尉之副，掌一部军务")
add_military("假司马",10, "军司马之副贰，佐领军务")

# ---- Level 9 郡守级 / 中郎将级 ----
add_civil("郡太守",  9, "一郡之长，治民进贤，秩二千石", adj={"req":{"glamourNeed":30}})
add_civil("郡长史",  9, "边郡设长史掌兵马，秩六百石")
add_civil("郡尉",    9, "大郡置尉分治诸县军事")
add_civil("盐官长",  9, "掌盐铁专卖，汉之重要财源")
add_civil("铁官长",  9, "掌铁冶铸器，汉之官营工官")
add_military("中郎将",9, "统领诸郎宿卫，秩比二千石")
add_military("骑都尉",9, "掌羽林骑兵，秩比二千石")
add_military("奉车都尉",9,"掌御乘舆车，秩比二千石")

# ---- Level 8 州佐 / 杂号将军 ----
add_civil("州治中从事",8, "州刺史首辅，居中治州事")
add_civil("州别驾从事",8, "州牧副手，出巡别乘传车")
add_civil("州主簿",    8, "掌州府文书机要")
add_civil("州功曹从事",8, "铨选州吏，掌州人事")
add_civil("典农都尉",  8, "掌屯田农事，汉末置")
add_military("度辽将军",8,"镇守辽东北境")
add_military("护羌校尉",8,"监护西羌诸部")
add_military("护乌桓校尉",8,"监护乌桓诸部")
add_military("使匈奴中郎将",8,"出使监护南匈奴")

# ---- Level 7 州牧级 / 四平将军 ----
add_civil("州刺史",  7, "一州监察，秩六百石，纠察郡国", adj={"req":{"glamourNeed":30}})
add_civil("州牧",    7, "一州军政长官，秩中二千石，汉末置", adj={"req":{"glamourNeed":45}})
add_civil("持节使",  7, "持节出使督察四方", adj={"req":{"glamourNeed":35}})
add_civil("刺史从事",7, "刺史属官，分管诸曹")
add_civil("都水长",  7, "掌水利河渠堤防")
add_military("平西将军",7,"平定西疆诸郡")
add_military("平东将军",7,"平定东方诸郡")
add_military("平南将军",7,"平定南土诸郡")
add_military("平北将军",7,"平定北疆诸郡")

# ---- Level 6 中央中阶 / 四镇将军 ----
add_civil("尚书仆射",6, "尚书令之副，参决政务机要，秩六百石", adj={"req":{"glamourNeed":25}})
add_civil("侍中",    6, "侍从天子左右，备顾问，秩比二千石", adj={"req":{"glamourNeed":45}})
add_civil("中常侍",  6, "内廷侍从，传达诏命，秩千石", adj={"req":{"glamourNeed":35}})
add_civil("给事中",  6, "随侍内朝，评驳奏章", adj={"req":{"glamourNeed":30}})
add_civil("御史中丞",6, "御史大夫之副，纠察百官，秩千石")
add_military("镇西将军",6,"镇守凉州及西域")
add_military("镇东将军",6,"镇守青徐扬诸郡")
add_military("镇南将军",6,"镇守荆襄江汉")
add_military("镇北将军",6,"镇守幽并河北")

# ---- Level 5 九卿级 / 四征将军 ----
add_civil("尚书令",  5, "尚书台之长，总揽政务，秩千石", adj={"req":{"glamourNeed":35}})
add_civil("光禄勋",  5, "掌宫殿宿卫及郎官选署，秩中二千石", adj={"req":{"glamourNeed":55}})
add_civil("太仆",    5, "掌天子车马及厩牧，秩中二千石")
add_civil("大鸿胪",  5, "掌诸侯王及四夷朝贡宾客，秩中二千石", adj={"req":{"glamourNeed":70}})
add_civil("大司农",  5, "掌全国财政钱谷赋税，秩中二千石")
add_civil("少府",    5, "掌皇室财政山海池泽之税，秩中二千石")
add_military("征西将军",5,"征讨西方之统帅")
add_military("征东将军",5,"征讨东方之统帅")
add_military("征南将军",5,"征讨南方之统帅")
add_military("征北将军",5,"征讨北方之统帅")

# ---- Level 4 上卿级 / 前后左右将军 ----
add_civil("太常",    4, "掌宗庙礼仪祭祀，九卿之首，秩中二千石", adj={"req":{"glamourNeed":55}})
add_civil("卫尉",    4, "掌宫门屯卫兵，秩中二千石")
add_civil("廷尉",    4, "掌刑狱律法，天下刑名之总，秩中二千石")
add_civil("宗正",    4, "掌皇族宗室属籍，秩中二千石", adj={"req":{"glamourNeed":45}})
add_civil("执金吾",  4, "掌京师巡察治安，秩中二千石", adj={"req":{"glamourNeed":40}})
add_civil("将作大匠",4, "掌宫室陵寝营造修缮，秩二千石")
add_military("前将军",4,"四军将军之首，主征伐")
add_military("后将军",4,"四军之殿，主镇守")
add_military("左将军",4,"四军之左翼统帅")
add_military("右将军",4,"四军之右翼统帅")

# ---- Level 3 三公级 / 重号将军 ----
add_civil("太尉",    3, "三公之首，掌四方兵事，秩万石", adj={"req":{"glamourNeed":50}})
add_civil("司徒",    3, "三公之二，掌教化民政，秩万石", adj={"req":{"glamourNeed":55}})
add_civil("司空",    3, "三公之三，掌水土工程，秩万石", adj={"req":{"glamourNeed":30}})
add_civil("司隶校尉",3, "督察三辅三河弘农，持节，秩比二千石")
add_civil("太子太傅",3, "太子之师，掌辅导太子，秩二千石", adj={"req":{"glamourNeed":50}})
add_military("车骑将军",3,"重号将军，统战车骑兵，秩万石")
add_military("骠骑将军",3,"重号将军，统骠骑劲旅，秩万石")
add_military("卫将军", 3,"重号将军，总领京师宿卫，秩万石")
add_military("护军将军",3,"掌监诸军，典武选，汉末置")

# ---- Level 2 上公级 / 顶级武将 ----
add_civil("太师",    2, "天子之师，上公之首，位在三公上", adj={"req":{"glamourNeed":65}})
add_civil("太傅",    2, "天子辅弼，上公之二，位在三公上", adj={"req":{"glamourNeed":70}})
add_civil("太保",    2, "保育天子，上公之三，位在三公上", adj={"req":{"glamourNeed":60}})
add_civil("丞相长史",2, "丞相府诸吏之长，参赞万机，秩千石")
add_civil("御史大夫",2, "副丞相，掌监察弹劾，秩万石")
add_military("大将军",  2, "最高武职，统天下兵马，位在三公上")
add_military("大司马",  2, "掌武事军事，冠于诸将军之上")
add_military("度辽大将军",2,"统北境诸军，镇守边疆")
add_military("辅国将军",2,"辅弼国政，汉末重号")

# ---- Level 1 相国级 / 府号将军 ----
add_civil("丞相",    1, "百官之长，总揽朝政，秩万石", adj={"req":{"glamourNeed":65}})
add_civil("相国",    1, "尊于丞相，权倾朝野，汉初置后复置", adj={"req":{"glamourNeed":70}})
add_civil("录尚书事",1, "总领尚书台，裁决万机，东汉末置", adj={"req":{"glamourNeed":50}})
add_civil("假节",    1, "持节督军，得斩违令者")
add_civil("光禄大夫",1, "顾问应对，秩比二千石，加官之重", adj={"req":{"glamourNeed":40}})
add_civil("谏议大夫",1, "掌议论讽谏，秩六百石", adj={"req":{"glamourNeed":35}})
add_military("骠骑大将军",1,"骠骑之极，位次丞相")
add_military("车骑大将军",1,"车骑之极，仪同三公")
add_military("卫大将军",  1,"宿卫之极，总领诸卫")
add_military("大将军(辅政)",1,"辅佐天子，权兼内外")

# ---- Level 0 至尊级 ----
add_civil("上公",    0, "人臣之极，位在三公之上", adj={"req":{"glamourNeed":70}})
add_civil("丞相(开府)",0,"开府置官属，总百揆", adj={"req":{"glamourNeed":65}})
add_civil("大司马(辅政)",0,"辅政大司马，权同三公", adj={"req":{"glamourNeed":55}})
add_civil("太傅(录尚书事)",0,"上公兼录尚书事，总揽万机", adj={"req":{"glamourNeed":70}})
add_military("大将军(录尚书事)",0,"大将军兼录尚书，军政大权一身")
add_military("大将军(都督中外)",0,"都督中外诸军事，天下兵马之帅")
add_military("大司马(都督中外)",0,"大司马都督中外，冠三军之首")

# ========== 写入 Excel ============
for r in rows:
    ws.append(r)

# ---- 样式 ----
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="000000")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

for c in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = bdr

# 隔级变色
light_blue = PatternFill("solid", fgColor="DCE6F1")
light_green = PatternFill("solid", fgColor="E2EFDA")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=21):
    level_val = row[4].value  # level 列
    fill = None
    if level_val is not None:
        if level_val % 2 == 0:
            fill = light_blue
        else:
            fill = light_green
    for cell in row:
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill

col_widths = [6,18,12,8,8,10,10,10,12,12,12,12,12,32,10,10,12,12,12,12]
for i,w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ---- 设计说明 Sheet ----
ws_note = wb.create_sheet("设计说明")
notes_data = {
    "A": [
        "三国官职数值设计说明",
        "",
        "一、设计约束",
        "1. 共15级官职（Level 14 最低 → Level 0 最高）",
        "2. 每级不少于8种官职（文官+武官）",
        "3. 增加属性最大值 ≤ 10",
        "4. 指挥(troopsLimit)范围 5000~15000",
        "5. 俸禄(cost)范围 20~320",
        "6. 功绩(meritNeeds)=晋升到下一级所需功绩，Level 0(最高)=0",
        "7. 功绩范围：1000(最低)~50000(最高)",
        "8. 需求等级(levelNeed)范围：5(最低)~50(最高)",
        "9. 数据从最低官职开始排序（Level 14 → Level 0）",
        "",
        "二、属性加成规则",
        "文官：偏重智力、政治、魅力加成",
        "武官：偏重统率、武力加成",
        "加成值随级别升高而递增",
        "",
        "三、属性需求规则",
        "文官：需求高智力、高政治",
        "武官：需求高统率、高武力",
        "需求值随级别升高而递增",
        "需求等级(levelNeed)：Level 14→5, Level 0→50",
    ],
    "C": [
        "四、级别体系概览",
        "Level 14：书佐亭长/伍长什长",
        "Level 13：功曹贼曹/队率屯长",
        "Level 12：县丞主簿/军候假候",
        "Level 11：县令县长/都尉别部司马",
        "Level 10：郡丞功曹/校尉军司马",
        "Level 9：郡守盐铁/中郎将骑都尉",
        "Level 8：州治中别驾/度辽护羌",
        "Level 7：刺史州牧/四平将军",
        "Level 6：尚书侍中/四镇将军",
        "Level 5：九卿级/四征将军",
        "Level 4：上卿级/前后左右将军",
        "Level 3：三公级/车骑骠骑卫将军",
        "Level 2：太师太傅/大将军大司马",
        "Level 1：丞相相国/府号大将军",
        "Level 0：上公太傅/大司马大将军",
        "",
        "所有官职名均出自两汉实际存在的官职体系，",
        "参考《汉书·百官公卿表》《后汉书·百官志》。",
        "",
        "五、数值公式（近似参考）",
        "troopsLimit ≈ 5000→15000，每级递增约+700",
        "cost ≈ 20→320，每级递增约+22",
        "meritNeeds(晋升下一级) ≈ 1000→50000，Level 0=0",
        "levelNeed ≈ 5→50 线性递增",
        "commandNeed(strength) ≈ 5 + (14-level) * 6",
        "politicsNeed(intel) ≈ 5 + (14-level) * 6.5",
    ],
}
for col, texts in notes_data.items():
    for i, t in enumerate(texts, start=1):
        cell = ws_note[f"{col}{i}"]
        cell.value = t
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 1:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        elif t.startswith("一、") or t.startswith("二、") or t.startswith("三、") or t.startswith("四、") or t.startswith("五、"):
            cell.font = Font(bold=True, size=11)
ws_note.column_dimensions["A"].width = 42
ws_note.column_dimensions["C"].width = 42

out = "d:/sango_infinity/Build/三国官职数值_v2.xlsx"
wb.save(out)
print(f"Done: {len(rows)} 官职 → {out}")
